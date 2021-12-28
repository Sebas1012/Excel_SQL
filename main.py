import json
import requests
from openpyxl import load_workbook

# Configuraciones Iniciales de openpyxl
workbook = load_workbook('./document/datos.xlsx')
sheet = workbook.active

# Lectura del archivo counter, el cual almacena la ultima fila que se itero para poder continuar desde la misma.
counter = open('./counter.txt', 'r')
i = int(counter.read())
j = 0

# Cabecereas para la peticion POST
headers = {'Content-Type':'application/json'}


for row in sheet.iter_rows(min_row=i, values_only=True):
  fecha_atencion = str(row[2])[:10]

  # Dicionario para ordenar los datos obtenidos de la lectura del archivo excel.
  body = {
          "tipo_doc" : row[0],
          "documento" : row[1],
          "fecha_atencion" : fecha_atencion,
          "tipo_atencion" : row[5],
          "medico" : row[6],
          "dx" : row[7],
          "descripcion_dx" : row[8],
          "lateralidad" : row[9],
          "av" : row[10],
          "tipo_av" : row[11],
          "emc" : row[12],
          "av_lb" : row[13],
          "observaciones" : row[14],
          "eps" : row[15],
        }

  body = json.dumps(body)

  request = requests.post('http://admin:test@127.0.0.1:4567/test', headers=headers, data=body)
  print(request)
    
  j+=1
  print(body)


j = j + i
j = str(j)

# Escritura en el archivo counter de la ultima fila iterada para poder contunuar desde la misma si hay actualizaciones en el archivo.
counter = open('./counter.txt', 'w')
counter.write(j)
counter.close()

    
    
