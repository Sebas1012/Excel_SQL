import json
import requests
from openpyxl import load_workbook

def read_xl(book, counter_file):
  book = load_workbook(str(book))
  sheet = book.active

  file = str(counter_file)

  file_path = open(file, 'r')
  i = int(file_path.read())
  j = 0

  for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    columns = row

  headers = {'Content-Type':'application/json'}

  for row in sheet.iter_rows(min_row=i, values_only=True):
    body = dict(zip(columns, row))

    body = json.dumps(body)

    request = requests.post('http://admin:test@127.0.0.1:4567/test', headers=headers, data=body)
    
    if request.status_code >= 200 and request.status_code < 300:
      print('Datos enviados con exito. Codigo: ' + str(request.status_code))
    else:
      print('Ocurrio un error. Codigo: ' + str(request.status_code))

    j+=1

  j = j + i
  j = str(j)

  file_path = open(file,'w')
  file_path.write(j)
  file_path.close


print(read_xl('./document/datos.xlsx', './counter.txt'))



